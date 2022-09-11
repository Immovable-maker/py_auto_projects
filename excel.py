import  openpyxl  as  op
path = r"C:\Users\Desktop\VS CODE\Project\practice\연습"
wb = op.load_workbook(path + "/test.xlsx")

ws = wb.active #활성화되어있는 시트 설정
print(ws)

ws = wb["USD"] #무 Sheet로 설정하기
print(ws)

ws_list = wb.sheetnames  #해당 Workbook의 시트 목록을 리스트로 저장
print(ws_list) #리스트 출력

#방법 1 : Sheet의 Cell 속성 사용하기
data1 = ws.cell(row=1, column=2).value

#방법 2 : 엑셀 인덱스(Range) 사용하기
data2 = ws["B1"].value

#위 결과 출력해보기
print("cell(1,2) : ", data1)
print('Range("B1"):', data2)

#"B1" Cell에 입력하기
ws.cell(row=1, column=2).value = "입력테스트1"

#"C1" Cell에 입력하기
ws["C1"].value = "입력테스트2"

wb.save("result.xlsx") #엑셀 파일 저장(파일명 : result.xlsx)